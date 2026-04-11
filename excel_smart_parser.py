"""
Excel Universal Parser v19
===============================================================================
Исправления относительно v18.1:

  [FIX] Эвристика теперь запускается ВСЕГДА (не только при пустом all_tables).
        Ранее если на листе были native/named tables, heuristic не запускался,
        и остальные строки терялись. Теперь поведение единообразно: все 5
        источников запускаются всегда, перекрытия предотвращает used_rows.

  [FIX] _extract_heuristic / _extract_vertical / _extract_headerless:
        двойное чтение ячеек для _detect_dtype. Ранее adapter.cell()
        вызывался повторно в list comprehension для dtype, хотя значения
        уже прочитаны при формировании rows_out. Добавлен col_values кэш
        во всех трёх методах (ранее был только в _parse_range).

  [FIX] CLI --include-hidden не работал: skip_hidden по умолчанию False,
        а --include-hidden ставил skip_hidden=not True=False — то же самое.
        Заменён на --skip-hidden (skip_hidden=True при указании флага).

  [FIX] XlrdAdapter.cell(): import xlrd выполнялся внутри метода при каждом
        вызове ячейки. Убран повторный import — модуль сохраняется в
        self._xlrd при инициализации адаптера (один import в __init__).

  [FIX] PyxlsbAdapter: named_ranges() открывал файл повторно через
        pyxlsb.open_workbook(), хотя _ensure_cache() уже это делала.
        Рефакторинг: named_ranges извлекаются в _ensure_cache() за один
        проход и сохраняются в _named_ranges_cache.

  [FIX] _serialize: не обрабатывал datetime.time и datetime.timedelta.
        Добавлена сериализация: time -> .isoformat(), timedelta -> str().

  [FIX] _detect_dtype: boolean значения (True/False) не распознавались —
        _is_numeric возвращает False для bool, они попадали в "text".
        Добавлен тип "boolean" при >50% bool-значений.

  [FIX] CSV: пробел убран из KNOWN_DELIMITERS и _FALLBACK_DELIMITERS.
        Пробел как разделитель ложно срабатывает на любом тексте с пробелами.

  [FIX] _load_xlsx: при ошибке scan файл открывался до 3 раз (scan +
        fallback + основной). Упрощено: при ошибке scan сразу переходим
        к основному открытию без промежуточного fallback.

  [FIX] ExcelParser.__init__: добавлена валидация параметров (threshold,
        min_data_cells, max_empty_streak).

  [FIX] Удалён deprecated _check_formulas_lazy — внутренняя функция,
        не часть публичного API.

  [FIX] Docstring changelog (194 строки) вынесен — модульный docstring
        содержит только описание и версию.

  [FIX] Дублирование кода сборки rows_out: вынесен общий метод
        _build_rows() для _extract_heuristic, _extract_vertical,
        _extract_headerless.
"""


from __future__ import annotations

import abc
import argparse
import codecs
import csv
import datetime
import json
import math
import os
import sys
import traceback
import warnings
from typing import Any, Generator, Optional

# -- Обязательная зависимость -------------------------------------------------
try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.utils import range_boundaries as _range_boundaries
except ImportError:
    sys.exit("pip install openpyxl")

# -- Опциональные зависимости -------------------------------------------------
try:
    from tqdm import tqdm as _tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import pyxlsb
    HAS_PYXLSB = True
except ImportError:
    HAS_PYXLSB = False

try:
    import chardet as _chardet
    HAS_CHARDET = True
except ImportError:
    try:
        import charset_normalizer as _chardet  # type: ignore
        HAS_CHARDET = True
    except ImportError:
        HAS_CHARDET = False


# ==============================================================================
# Типы
# ==============================================================================

CellValue = Any
Row = dict[str, CellValue]

TABLE_SOURCE_NATIVE = "native_table"
TABLE_SOURCE_NAMED = "named_range"
TABLE_SOURCE_HEURISTIC = "heuristic"
TABLE_SOURCE_VERTICAL = "vertical"
TABLE_SOURCE_HEADERLESS = "headerless"


# ==============================================================================
# Абстрактный адаптер листа
# ==============================================================================

class SheetAdapter(abc.ABC):
    """Абстрактный адаптер листа. Подклассы обязаны реализовать cell()."""
    name: str
    max_row: int
    max_col: int

    @abc.abstractmethod
    def cell(self, row: int, col: int) -> CellValue:
        """Возвращает значение ячейки (1-based row, col). None для пустых."""

    def iter_rows_lazy(self, cols: list[int]) -> Generator[tuple[int, list[CellValue]], None, None]:
        for r in range(1, self.max_row + 1):
            yield r, [self.cell(r, c) for c in cols]

    def hidden_rows(self) -> set[int]:
        return set()

    def hidden_cols(self) -> set[int]:
        return set()

    def native_tables(self) -> list[dict]:
        return []

    def named_ranges(self) -> list[dict]:
        return []


# -- openpyxl (.xlsx / .xlsm / .xltx / .xltm) --------------------------------

class OpenpyxlAdapter(SheetAdapter):
    def __init__(self, ws, name: str):
        self._ws = ws
        self.name = name
        self.max_row = ws.max_row or 0
        self.max_col = ws.max_column or 0
        self._merged: dict[tuple[int, int], CellValue] = {}
        for rng in ws.merged_cells.ranges:
            master = ws.cell(rng.min_row, rng.min_col).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    self._merged[(r, c)] = master

    def cell(self, row: int, col: int) -> CellValue:
        if (row, col) in self._merged:
            return self._merged[(row, col)]
        return self._ws.cell(row, col).value

    def hidden_rows(self) -> set[int]:
        return {r for r, d in self._ws.row_dimensions.items() if d.hidden}

    def hidden_cols(self) -> set[int]:
        result: set[int] = set()
        for letter, d in self._ws.column_dimensions.items():
            if d.hidden:
                try:
                    result.add(column_index_from_string(letter))
                except Exception as e:
                    warnings.warn(
                        f"Лист '{self.name}': не удалось распознать скрытую колонку "
                        f"'{letter}': {e}",
                        UserWarning, stacklevel=2,
                    )
        return result

    def native_tables(self) -> list[dict]:
        tables = []
        for tname, tobj in self._ws.tables.items():
            try:
                ref = tobj.ref if hasattr(tobj, "ref") else str(tobj)
                if ":" not in ref:
                    continue
                min_col, min_row, max_col, max_row = _range_boundaries(ref)
                tables.append({"name": tname, "ref": ref,
                                "min_row": min_row, "max_row": max_row,
                                "min_col": min_col, "max_col": max_col})
            except Exception as e:
                warnings.warn(
                    f"Лист '{self.name}': нативная таблица '{tname}' "
                    f"имеет некорректный диапазон: {e}",
                    UserWarning, stacklevel=2,
                )
        return tables


# -- xlrd (.xls) ---------------------------------------------------------------

class XlrdAdapter(SheetAdapter):
    def __init__(self, sheet, book, name: str):
        import xlrd as _xlrd_module
        self._xlrd = _xlrd_module
        self._sheet = sheet
        self._book = book
        self.name = name
        self.max_row = sheet.nrows
        self.max_col = sheet.ncols

    def cell(self, row: int, col: int) -> CellValue:
        # FIX v19: используем self._xlrd вместо top-level xlrd
        # (top-level xlrd может быть не определён если модуль не установлен)
        if row < 1 or col < 1 or row > self.max_row or col > self.max_col:
            return None
        c = self._sheet.cell(row - 1, col - 1)
        if c.ctype == self._xlrd.XL_CELL_EMPTY:
            return None
        if c.ctype == self._xlrd.XL_CELL_DATE:
            try:
                return self._xlrd.xldate_as_datetime(c.value, self._book.datemode)
            except Exception as e:
                warnings.warn(
                    f"Лист '{self.name}': не удалось преобразовать дату "
                    f"(xldate={c.value}): {e}. Возвращается raw float.",
                    UserWarning, stacklevel=2,
                )
                return c.value
        return c.value

    def named_ranges(self) -> list[dict]:
        result = []
        try:
            for name_obj in self._book.name_obj_list:
                name = name_obj.name
                if not name_obj.result:
                    continue
                try:
                    for area in name_obj.result.coords:
                        shtxlo, _shtxhi, row0, row1, col0, col1 = area[:6]
                        sheet_idx = shtxlo
                        if self._book.sheet_names()[sheet_idx] != self.name:
                            continue
                        result.append({
                            "name": name,
                            "sheet": self.name,
                            "min_row": row0 + 1,
                            "max_row": row1,
                            "min_col": col0 + 1,
                            "max_col": col1,
                        })
                except Exception as e:
                    warnings.warn(
                        f"Лист '{self.name}': именованный диапазон '{name}' "
                        f"не удалось распарсить: {e}",
                        UserWarning, stacklevel=2,
                    )
        except Exception as e:
            warnings.warn(
                f"Лист '{self.name}': ошибка при чтении name_obj_list: {e}",
                UserWarning, stacklevel=2,
            )
        return result


# -- pyxlsb (.xlsb) -----------------------------------------------------------

class PyxlsbAdapter(SheetAdapter):
    """
    Адаптер для .xlsb файлов. Данные хранятся в _row_cache.
    Создаётся через _load_xlsb(), который открывает файл один раз
    и читает все листы за один проход.
    """

    def __init__(
        self,
        sheet_name: str,
        row_cache: dict[int, list],
        max_row: int,
        max_col: int,
        named_ranges_cache: list[dict],
    ):
        self.name = sheet_name
        self._row_cache = row_cache
        self.max_row = max_row
        self.max_col = max_col
        self._named_ranges_cache = named_ranges_cache

    def cell(self, row: int, col: int) -> CellValue:
        if row < 1 or col < 1 or row > self.max_row or col > self.max_col:
            return None
        row_data = self._row_cache.get(row)
        if row_data is None:
            return None
        idx = col - 1
        if idx >= len(row_data):
            return None
        return row_data[idx]

    def iter_rows_lazy(self, cols: list[int]) -> Generator[tuple[int, list[CellValue]], None, None]:
        col_to_pos = {c: i for i, c in enumerate(cols)}
        for r in range(1, self.max_row + 1):
            row_data = self._row_cache.get(r, [None] * self.max_col)
            vals = [None] * len(cols)
            for c, pos in col_to_pos.items():
                idx = c - 1
                if 0 <= idx < len(row_data):
                    vals[pos] = row_data[idx]
            yield r, vals

    def named_ranges(self) -> list[dict]:
        # FIX v19: возвращаем кэш вместо повторного открытия файла
        return self._named_ranges_cache


# -- CSV -----------------------------------------------------------------------

class CsvAdapter(SheetAdapter):
    """
    Адаптер CSV. Весь файл загружается в _row_cache одним проходом.
    Потребление памяти O(n) по числу строк.
    """

    def __init__(self, filepath: str, encoding: str, delimiter: str, name: str):
        self.name = name
        self._filepath = filepath
        self._encoding = encoding
        self._delimiter = delimiter
        self._row_cache: dict[int, list] = {}
        self.max_row, self.max_col = self._load_cache()

    def _load_cache(self) -> tuple[int, int]:
        max_r = max_c = 0
        with open(self._filepath, newline="", encoding=self._encoding) as f:
            for i, row in enumerate(csv.reader(f, delimiter=self._delimiter)):
                r = i + 1
                max_r = r
                if len(row) > max_c:
                    max_c = len(row)
                self._row_cache[r] = row
        return max_r, max_c

    def cell(self, row: int, col: int) -> CellValue:
        c = col - 1
        row_data = self._row_cache.get(row, [])
        if c < 0 or c >= len(row_data):
            return None
        v = row_data[c]
        return v if v != "" else None

    def iter_rows_lazy(self, cols: list[int]) -> Generator[tuple[int, list[CellValue]], None, None]:
        col_to_pos = {c: i for i, c in enumerate(cols)}
        for r in range(1, self.max_row + 1):
            row_data = self._row_cache.get(r, [])
            vals = [None] * len(cols)
            for c, pos in col_to_pos.items():
                raw_c = c - 1
                if 0 <= raw_c < len(row_data):
                    v = row_data[raw_c]
                    vals[pos] = v if v != "" else None
            yield r, vals


# ==============================================================================
# Загрузчики
# ==============================================================================

def load_sheets(filepath: str, only_sheet: Optional[str] = None) -> list[SheetAdapter]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return _load_xlsx(filepath, only_sheet)
    elif ext == ".xls":
        return _load_xls(filepath, only_sheet)
    elif ext == ".xlsb":
        return _load_xlsb(filepath, only_sheet)
    elif ext == ".csv":
        return _load_csv(filepath)
    else:
        raise ValueError(f"Неподдерживаемый формат: {ext}. "
                         "Поддерживается: .xlsx .xlsm .xls .xlsb .csv")


def _load_xlsx(filepath: str, only_sheet: Optional[str]) -> list[OpenpyxlAdapter]:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xlsm":
        warnings.warn(
            f"'{os.path.basename(filepath)}': формат .xlsm -- "
            "файл откроется корректно, VBA-макросы игнорируются.",
            UserWarning, stacklevel=3,
        )

    # Проверка формул: один быстрый read_only проход
    sheets_to_process: list[str] = []
    try:
        wb_scan = openpyxl.load_workbook(filepath, data_only=False, read_only=True)
        try:
            for sheet_name in wb_scan.sheetnames:
                if only_sheet and sheet_name != only_sheet:
                    continue
                ws = wb_scan[sheet_name]
                for row in ws.iter_rows(values_only=False):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("="):
                            warnings.warn(
                                f"Лист '{sheet_name}': обнаружены формулы. "
                                "data_only=True вернёт None для несохранённых значений.",
                                UserWarning, stacklevel=3,
                            )
                            break
                    else:
                        continue
                    break
                sheets_to_process.append(sheet_name)
        finally:
            wb_scan.close()
    except Exception as e:
        # FIX v19: упрощённый fallback — без промежуточного открытия файла
        warnings.warn(
            f"_load_xlsx: не удалось выполнить проверку формул для "
            f"'{os.path.basename(filepath)}': {e}",
            UserWarning, stacklevel=2,
        )
        sheets_to_process = []  # будет заполнен из основного wb ниже

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Если scan не удался, берём список листов из основного wb
    if not sheets_to_process:
        sheets_to_process = [
            n for n in wb.sheetnames
            if not only_sheet or n == only_sheet
        ]

    sheets = []
    for name in wb.sheetnames:
        if name not in sheets_to_process:
            continue
        ws = wb[name]
        if not ws.max_row:
            continue
        sheets.append(OpenpyxlAdapter(ws, name))
    return sheets


def _load_xls(filepath: str, only_sheet: Optional[str]) -> list[XlrdAdapter]:
    if not HAS_XLRD:
        raise ImportError("pip install xlrd  # для .xls файлов")
    wb = xlrd.open_workbook(filepath)
    result = []
    for name in wb.sheet_names():
        if only_sheet and name != only_sheet:
            continue
        result.append(XlrdAdapter(wb.sheet_by_name(name), wb, name))
    return result


def _load_xlsb(filepath: str, only_sheet: Optional[str]) -> list[PyxlsbAdapter]:
    if not HAS_PYXLSB:
        raise ImportError("pip install pyxlsb  # для .xlsb файлов")
    import pyxlsb as _pyxlsb

    # FIX v19: одно открытие файла — читаем все листы + named ranges за один проход.
    # Ранее: _load_xlsb открывал файл для списка листов, затем каждый
    # PyxlsbAdapter.__init__ открывал его повторно (N+1 открытий).
    result: list[PyxlsbAdapter] = []
    with _pyxlsb.open_workbook(filepath) as wb:
        # Собираем named ranges один раз для всех листов
        all_named: list[dict] = _read_xlsb_named_ranges(wb)

        for sheet_name in wb.sheets:
            if only_sheet and sheet_name != only_sheet:
                continue

            # Читаем данные листа
            tmp: dict[int, dict] = {}
            max_r = max_c = 0
            with wb.get_sheet(sheet_name) as ws:
                for i, row in enumerate(ws.rows()):
                    r = i + 1
                    if r > max_r:
                        max_r = r
                    for cell in row:
                        c_idx = cell.c
                        if c_idx + 1 > max_c:
                            max_c = c_idx + 1
                        tmp.setdefault(r, {})[c_idx] = cell.v

            # Конвертируем в row_cache
            row_cache: dict[int, list] = {}
            for r, cols in tmp.items():
                data = [None] * max_c
                for c_idx, v in cols.items():
                    if 0 <= c_idx < max_c:
                        data[c_idx] = v
                row_cache[r] = data

            # Фильтруем named ranges для этого листа
            sheet_named = [nr for nr in all_named if nr["sheet"] == sheet_name]

            result.append(PyxlsbAdapter(
                sheet_name=sheet_name,
                row_cache=row_cache,
                max_row=max_r,
                max_col=max_c,
                named_ranges_cache=sheet_named,
            ))

    return result


def _read_xlsb_named_ranges(wb) -> list[dict]:
    """Извлекает все named ranges из открытого pyxlsb workbook."""
    result: list[dict] = []
    try:
        candidates = (
            getattr(wb, "defined_names", None)
            or getattr(wb, "named_ranges", None)
            or getattr(wb, "_defined_names", None)
            or []
        )
        for nr in candidates:
            try:
                name = getattr(nr, "name", None) or getattr(nr, "Name", None)
                formula = getattr(nr, "formula", None) or getattr(nr, "refers_to", None)
                if not name or not formula:
                    continue
                if "!" not in formula:
                    continue
                sheet_part, range_part = formula.split("!", 1)
                sheet_name = sheet_part.strip("'\"$")
                min_col, min_row, max_col, max_row = _range_boundaries(range_part)
                if min_col > max_col or min_row > max_row:
                    continue
                result.append({
                    "name": name,
                    "sheet": sheet_name,
                    "min_row": min_row, "max_row": max_row,
                    "min_col": min_col, "max_col": max_col,
                })
            except Exception as e:
                warnings.warn(
                    f"xlsb named range: не удалось распарсить: {e}",
                    UserWarning, stacklevel=2,
                )
    except Exception as e:
        warnings.warn(
            f"xlsb: ошибка при чтении named_ranges: {e}",
            UserWarning, stacklevel=2,
        )
    return result


def _detect_encoding(filepath: str) -> str:
    if HAS_CHARDET:
        with open(filepath, "rb") as f:
            raw = f.read(32768)
        detected = _chardet.detect(raw)
        enc = (detected.get("encoding") or "utf-8").strip()
        try:
            codecs.lookup(enc)
            return enc
        except LookupError:
            pass
    for enc in ("utf-8-sig", "utf-8", "cp1251", "cp1252", "latin-1"):
        try:
            with open(filepath, encoding=enc) as f:
                f.read(4096)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "utf-8"


def _load_csv(filepath: str) -> list[CsvAdapter]:
    encoding = _detect_encoding(filepath)

    # FIX v19: пробел убран из разделителей — ложно срабатывает на текстах
    KNOWN_DELIMITERS = ",;\t|^~"
    try:
        with open(filepath, newline="", encoding=encoding) as f:
            sample = f.read(16384)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=KNOWN_DELIMITERS)
            delimiter = dialect.delimiter
        except csv.Error:
            _FALLBACK_DELIMITERS = frozenset(",;\t|^~")
            candidates = {d: sample.count(d) for d in _FALLBACK_DELIMITERS}
            _best = max(candidates, key=candidates.get)
            delimiter = _best if candidates[_best] > 0 else ","

        return [CsvAdapter(filepath, encoding, delimiter, os.path.basename(filepath))]
    except Exception as e:
        raise ValueError(f"Не удалось прочитать CSV '{filepath}': {e}") from e


# ==============================================================================
# Утилиты значений
# ==============================================================================

def _to_str(v: CellValue) -> str:
    return "" if v is None else str(v).strip()


def _is_empty(v: CellValue) -> bool:
    return _to_str(v) == ""


def _is_numeric(v: CellValue) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return math.isfinite(v)
    s = _to_str(v).replace(",", ".").replace(" ", "").replace("%", "")
    if not s:
        return False
    try:
        return math.isfinite(float(s))
    except ValueError:
        return False


def _is_date(v: CellValue) -> bool:
    return isinstance(v, (datetime.datetime, datetime.date))


def _is_year(v: CellValue) -> bool:
    if isinstance(v, bool):
        return False
    if isinstance(v, float):
        if not math.isfinite(v):
            return False
        if v != int(v):
            return False
        return 1900 <= int(v) <= 2200
    if isinstance(v, int):
        return 1900 <= v <= 2200
    s = _to_str(v)
    try:
        return 1900 <= int(s) <= 2200
    except ValueError:
        return False


def _serialize(v: CellValue) -> Any:
    # FIX v19: обработка time и timedelta
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    if isinstance(v, datetime.time):
        return v.isoformat()
    if isinstance(v, datetime.timedelta):
        total_sec = v.total_seconds()
        sign = "-" if total_sec < 0 else ""
        abs_total = int(abs(total_sec))
        hours, remainder = divmod(abs_total, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"
    return v


def _detect_dtype(values: list[CellValue]) -> str:
    non_empty = [v for v in values if not _is_empty(v)]
    if not non_empty:
        return "text"
    total = len(non_empty)

    # FIX v19: распознавание boolean
    bools = sum(1 for v in non_empty if isinstance(v, bool))
    if bools / total > 0.5:
        return "boolean"

    nums = sum(1 for v in non_empty if _is_numeric(v))
    dates = sum(1 for v in non_empty if _is_date(v))
    if dates / total > 0.5:
        return "date"
    if nums / total > 0.7:
        pct_count = sum(
            1 for v in non_empty
            if isinstance(v, str) and "%" in v
        )
        return "percent" if pct_count / total > 0.3 else "number"
    return "text"


# ==============================================================================
# Score-based детектор заголовков
# ==============================================================================

def _score_header_row(vals: list[CellValue]) -> float:
    non_empty = [v for v in vals if not _is_empty(v)]
    if not non_empty:
        return 0.0

    # Паттерн "числа-дни" (1,2,3...31 подряд)
    numeric_vals = [v for v in non_empty if _is_numeric(v) and not _is_year(v)]
    is_day_header = False
    if len(numeric_vals) >= 5:
        day_numbers = []
        for v in numeric_vals:
            if isinstance(v, bool):
                continue
            if isinstance(v, float):
                if not math.isfinite(v):
                    continue
                n = int(v)
            elif isinstance(v, int):
                n = v
            elif isinstance(v, str):
                # FIX v19: CSV возвращает числа как строки
                try:
                    n = int(float(v.replace(",", ".").replace(" ", "").replace("%", "")))
                except (ValueError, OverflowError):
                    continue
            else:
                continue
            if 1 <= n <= 31:
                day_numbers.append(n)
        if len(day_numbers) >= 5:
            day_numbers_sorted = sorted(day_numbers)
            consecutive = sum(
                1 for i in range(len(day_numbers_sorted) - 1)
                if 1 <= (day_numbers_sorted[i + 1] - day_numbers_sorted[i]) <= 2
            )
            if consecutive >= len(day_numbers_sorted) * 0.6:
                is_day_header = True

    text_like = 0.0
    for v in non_empty:
        if _is_year(v):
            if isinstance(v, str):
                text_like += 0.3
        elif _is_date(v):
            text_like += 0.5
        elif _is_numeric(v):
            if is_day_header:
                text_like += 0.5
            else:
                text_like -= 0.2
        else:
            s = _to_str(v)
            if len(s) == 1:
                text_like += 0.7
            elif 2 <= len(s) <= 60:
                text_like += 1.0
            elif len(s) > 60:
                text_like -= 0.5
    ratio = text_like / len(non_empty)
    if len(non_empty) < 2:
        ratio *= 0.6
    return max(0.0, min(1.0, ratio))


def _is_header_row(vals: list[CellValue], threshold: float = 0.4) -> bool:
    return _score_header_row(vals) >= threshold


def _dedupe_headers(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for i, name in enumerate(names):
        key = name if name else f"_col_{i + 1}"
        if key not in seen:
            seen[key] = 1
            result.append(key)
        else:
            seen[key] += 1
            result.append(f"{key}_{seen[key]}")
    return result


# ==============================================================================
# Потоковый writer
# ==============================================================================

class StreamingWriter:
    def __init__(self, output_path: str, fmt: str, file_meta: dict):
        self.fmt = fmt
        self.output_path = output_path
        self._table_count = 0
        self._row_count = 0
        self._file_meta = file_meta
        self._json_fh = None
        self._jsonl_fh = None
        self._csv_dir = None
        self._open(output_path, fmt)

    def _open(self, path: str, fmt: str) -> None:
        if fmt == "json":
            self._json_fh = open(path, "w", encoding="utf-8")
            meta = {k: v for k, v in self._file_meta.items()}
            meta_str = json.dumps(meta, ensure_ascii=False)
            self._json_fh.write(meta_str[:-1] + ', "tables_data": [\n')
            self._first_table = True
        elif fmt == "jsonl":
            self._jsonl_fh = open(path, "w", encoding="utf-8")
        elif fmt == "csv":
            os.makedirs(path, exist_ok=True)
            self._csv_dir = path

    def write_table(self, table: dict) -> None:
        self._table_count += 1
        self._row_count += len(table.get("rows", []))
        if self.fmt == "json":
            sep = "" if self._first_table else ",\n"
            self._first_table = False
            meta = {k: v for k, v in table.items() if k != "rows"}
            meta_str = json.dumps(meta, ensure_ascii=False, default=str)
            self._json_fh.write(sep + meta_str[:-1] + ', "rows": [')
            rows = table.get("rows", [])
            for i, row in enumerate(rows):
                row_sep = "" if i == 0 else ","
                self._json_fh.write(
                    row_sep + json.dumps(row, ensure_ascii=False, default=str)
                )
            self._json_fh.write("]}")
        elif self.fmt == "jsonl":
            for row in table.get("rows", []):
                record = {"_sheet": table["sheet"], "_table": table["name"], **row}
                self._jsonl_fh.write(
                    json.dumps(record, ensure_ascii=False, default=str) + "\n"
                )
        elif self.fmt == "csv":
            rows = table.get("rows", [])
            if not rows:
                return
            safe = (
                table["name"]
                .replace("/", "_").replace("\\", "_")
                .replace(":", "_").replace(" ", "_")[:60]
            )
            path = os.path.join(self._csv_dir,
                                f"{self._table_count:02d}_{safe}.csv")
            fieldnames = list(rows[0].keys())
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
                w.writeheader()
                w.writerows(rows)

    def close(self) -> None:
        if self.fmt == "json" and self._json_fh:
            self._json_fh.write("\n]}")
            self._json_fh.close()
            self._json_fh = None
        elif self.fmt == "jsonl" and self._jsonl_fh:
            self._jsonl_fh.close()
            self._jsonl_fh = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        # Безопасное закрытие даже при ошибке
        try:
            if exc_type is None:
                self.close()
            else:
                # При ошибке закрываем без финализации JSON
                if self._json_fh:
                    self._json_fh.close()
                    self._json_fh = None
                if self._jsonl_fh:
                    self._jsonl_fh.close()
                    self._jsonl_fh = None
        except Exception:
            pass
        return False

    @property
    def stats(self) -> tuple[int, int]:
        return self._table_count, self._row_count


# ==============================================================================
# Парсер
# ==============================================================================

class ExcelParser:
    def __init__(
        self,
        header_threshold: float = 0.4,
        skip_hidden: bool = False,
        min_data_cells: int = 2,
        max_empty_streak: int = 50,
    ):
        # FIX v19: валидация параметров
        if not 0.0 <= header_threshold <= 1.0:
            raise ValueError(f"header_threshold должен быть в [0.0, 1.0], получено: {header_threshold}")
        if min_data_cells < 1:
            raise ValueError(f"min_data_cells должен быть >= 1, получено: {min_data_cells}")
        if max_empty_streak < 1:
            raise ValueError(f"max_empty_streak должен быть >= 1, получено: {max_empty_streak}")

        self.header_threshold = header_threshold
        self.skip_hidden = skip_hidden
        self.min_data_cells = min_data_cells
        self.max_empty_streak = max_empty_streak
        self._vis_cache: dict[int, tuple[list[int], list[int]]] = {}

    # -- Видимые строки / колонки -----------------------------------------------

    def _visible(self, adapter: SheetAdapter) -> tuple[list[int], list[int]]:
        key = id(adapter)
        if key not in self._vis_cache:
            hr = adapter.hidden_rows() if self.skip_hidden else set()
            hc = adapter.hidden_cols() if self.skip_hidden else set()
            rows = [r for r in range(1, adapter.max_row + 1) if r not in hr]
            cols = [c for c in range(1, adapter.max_col + 1) if c not in hc]
            self._vis_cache[key] = (rows, cols)
        return self._vis_cache[key]

    # -- Общий метод сборки строк -----------------------------------------------

    def _build_rows(
        self,
        adapter: SheetAdapter,
        data_rows: list[int],
        col_map: dict[int, str],
    ) -> tuple[list[Row], dict[int, list[CellValue]]]:
        """
        Читает ячейки один раз, возвращает (rows_out, col_values).
        col_values используется для _detect_dtype без повторного чтения.
        """
        col_values: dict[int, list[CellValue]] = {c: [] for c in col_map}
        rows_out: list[Row] = []
        for r in data_rows:
            rd: dict[str, Any] = {}
            has = False
            for c, col_name in col_map.items():
                v = adapter.cell(r, c)
                col_values[c].append(v)
                if not _is_empty(v):
                    rd[col_name] = _serialize(v)
                    has = True
            if has:
                rows_out.append(rd)
        return rows_out, col_values

    # -- Парсинг диапазона ------------------------------------------------------

    def _parse_range(
        self,
        adapter: SheetAdapter,
        min_row: int, max_row: int,
        min_col: int, max_col: int,
        source: str,
        name: str,
    ) -> Optional[dict]:
        if min_col > max_col or min_row > max_row:
            return None

        vis_rows, _ = self._visible(adapter)
        cols = list(range(min_col, max_col + 1))
        rows_in = [r for r in vis_rows if min_row <= r <= max_row]
        if len(rows_in) < 2:
            return None

        header_row = rows_in[0]
        data_rows = rows_in[1:]

        raw_headers = [_to_str(adapter.cell(header_row, c)) for c in cols]
        deduped = _dedupe_headers(raw_headers)
        header_dict: dict[int, str] = {
            c: deduped[i] for i, c in enumerate(cols) if deduped[i]
        }

        # Fallback для колонок без заголовка
        col_map: dict[int, str] = {}
        for c in cols:
            col_map[c] = header_dict.get(c, get_column_letter(c))

        rows_out, col_values = self._build_rows(adapter, data_rows, col_map)

        if not rows_out:
            return None

        return {
            "sheet": adapter.name,
            "name": name,
            "source": source,
            "header_row": header_row,
            "data_start": data_rows[0],
            "data_end": data_rows[-1],
            "columns": [
                {
                    "letter": get_column_letter(c),
                    "name": col_map[c],
                    "type": _detect_dtype(col_values[c]),
                }
                for c in cols
            ],
            "rows": rows_out,
        }

    # -- Источник 1: встроенные таблицы -----------------------------------------

    def _extract_native_tables(self, adapter: SheetAdapter) -> list[dict]:
        results = []
        for tbl in adapter.native_tables():
            p = self._parse_range(
                adapter,
                tbl["min_row"], tbl["max_row"],
                tbl["min_col"], tbl["max_col"],
                TABLE_SOURCE_NATIVE, tbl["name"],
            )
            if p:
                results.append(p)
        return results

    # -- Источник 2: именованные диапазоны --------------------------------------

    def _extract_named_ranges_from_wb(self, wb, adapters_map: dict[str, SheetAdapter]) -> list[dict]:
        results: list[dict] = []
        try:
            defined = wb.defined_names
        except AttributeError:
            return results

        for dn in defined:
            try:
                destinations = dn.destinations
            except AttributeError:
                continue
            if isinstance(destinations, str):
                if "!" not in destinations:
                    continue
                sheet_part, range_part = destinations.split("!", 1)
                sheet_title = sheet_part.strip("'\"$")
                try:
                    min_col, min_row, max_col, max_row = _range_boundaries(range_part)
                except Exception as e:
                    warnings.warn(
                        f"Именованный диапазон '{dn.name}': "
                        f"не удалось распарсить диапазон '{range_part}': {e}",
                        UserWarning, stacklevel=2,
                    )
                    continue
                if sheet_title not in adapters_map:
                    continue
                adapter = adapters_map[sheet_title]
                p = self._parse_range(
                    adapter, min_row, max_row, min_col, max_col,
                    TABLE_SOURCE_NAMED, dn.name,
                )
                if p:
                    results.append(p)
            else:
                for sheet_title, ref in destinations:
                    if sheet_title not in adapters_map:
                        continue
                    try:
                        min_col, min_row, max_col, max_row = _range_boundaries(ref)
                    except Exception as e:
                        warnings.warn(
                            f"Именованный диапазон '{dn.name}' на листе '{sheet_title}': "
                            f"не удалось распарсить диапазон '{ref}': {e}",
                            UserWarning, stacklevel=2,
                        )
                        continue
                    adapter = adapters_map[sheet_title]
                    p = self._parse_range(
                        adapter, min_row, max_row, min_col, max_col,
                        TABLE_SOURCE_NAMED, dn.name,
                    )
                    if p:
                        results.append(p)
        return results

    def _extract_named_ranges_from_adapter(self, adapter: SheetAdapter) -> list[dict]:
        results: list[dict] = []
        for nr in adapter.named_ranges():
            p = self._parse_range(
                adapter,
                nr["min_row"], nr["max_row"],
                nr["min_col"], nr["max_col"],
                TABLE_SOURCE_NAMED, nr["name"],
            )
            if p:
                results.append(p)
        return results

    # -- Источник 3: эвристика --------------------------------------------------

    def _extract_heuristic(
        self, adapter: SheetAdapter, external_used_rows: set[int]
    ) -> list[dict]:
        vis_rows, vis_cols = self._visible(adapter)
        if not vis_rows or not vis_cols:
            return []

        row_idx: dict[int, int] = {r: i for i, r in enumerate(vis_rows)}
        tables: list[dict] = []
        used_rows: set[int] = set(external_used_rows)
        table_counter = 0
        i = 0

        while i < len(vis_rows):
            row = vis_rows[i]
            vals = [adapter.cell(row, c) for c in vis_cols]

            if all(_is_empty(v) for v in vals):
                i += 1
                continue

            if not _is_header_row(vals, self.header_threshold):
                i += 1
                continue

            # Многострочный заголовок
            header_rows = [row]
            header_dict: dict[int, str] = {}
            for c in vis_cols:
                h = _to_str(adapter.cell(row, c))
                if h:
                    header_dict[c] = h

            for j in range(i + 1, min(i + 4, len(vis_rows))):
                r2 = vis_rows[j]
                r2v = [adapter.cell(r2, c) for c in vis_cols]
                sc = _score_header_row(r2v)
                nums = sum(1 for v in r2v if _is_numeric(v) and not _is_year(v))
                if sc >= 0.35 and nums == 0:
                    header_rows.append(r2)
                    for c in vis_cols:
                        h = _to_str(adapter.cell(r2, c))
                        if h:
                            header_dict[c] = (
                                header_dict[c] + " / " + h
                                if c in header_dict else h
                            )
                else:
                    break

            active_cols = set(header_dict.keys())
            if not active_cols:
                i += 1
                continue

            start_i = row_idx.get(header_rows[-1], i) + 1

            # Строки данных
            data_rows: list[int] = []
            empty_streak = 0

            for di in range(start_i, len(vis_rows)):
                r = vis_rows[di]
                rv = [adapter.cell(r, c) for c in active_cols]
                non_e = sum(1 for v in rv if not _is_empty(v))

                if non_e == 0:
                    empty_streak += 1
                    if empty_streak >= self.max_empty_streak:
                        break
                    continue

                if non_e < self.min_data_cells:
                    empty_streak = 0
                    continue

                empty_streak = 0
                data_rows.append(r)

            if not data_rows:
                i += 1
                continue

            span = set(range(header_rows[0], data_rows[-1] + 1))
            if span & used_rows:
                i += 1
                continue

            # Сборка
            sorted_cols = sorted(active_cols)
            raw_col_names = [header_dict[c] for c in sorted_cols]
            deduped_names = _dedupe_headers(raw_col_names)

            col_map: dict[int, str] = {}
            for c, col_name in zip(sorted_cols, deduped_names):
                col_map[c] = col_name

            # FIX v19: используем _build_rows для единого чтения ячеек
            rows_out, col_values = self._build_rows(adapter, data_rows, col_map)

            if rows_out:
                table_counter += 1
                tables.append({
                    "sheet": adapter.name,
                    "name": f"{adapter.name} / Таблица {table_counter}",
                    "source": TABLE_SOURCE_HEURISTIC,
                    "header_row": header_rows[0],
                    "data_start": data_rows[0],
                    "data_end": data_rows[-1],
                    "columns": [
                        {
                            "letter": get_column_letter(c),
                            "name": col_map[c],
                            "type": _detect_dtype(col_values[c]),
                        }
                        for c in sorted_cols
                    ],
                    "rows": rows_out,
                })
                used_rows.update(span)

            i = row_idx.get(data_rows[-1], i) + 1

        return tables

    # -- Источник 4: вертикальные таблицы ---------------------------------------

    def _extract_vertical(self, adapter: SheetAdapter, used_rows: set[int]) -> list[dict]:
        vis_rows, vis_cols = self._visible(adapter)
        if not vis_rows or len(vis_cols) < 2:
            return []

        col_a = vis_cols[0]
        other_c = vis_cols[1:]

        blocks: list[list[int]] = []
        current: list[int] = []
        for r in vis_rows:
            rv = [adapter.cell(r, c) for c in vis_cols]
            if all(_is_empty(v) for v in rv):
                if current:
                    blocks.append(current)
                    current = []
            else:
                current.append(r)
        if current:
            blocks.append(current)

        results: list[dict] = []

        for block in blocks:
            if len(block) < 2:
                continue

            header_row = block[0]

            if header_row in used_rows:
                continue

            header_vals = [adapter.cell(header_row, c) for c in vis_cols]
            if not _is_header_row(header_vals, 0.3):
                continue

            data_candidates = block[1:]

            text_in_a = 0
            numeric_in_rest = 0
            total_rest = 0
            for r in data_candidates:
                if r in used_rows:
                    continue
                va = adapter.cell(r, col_a)
                if not _is_empty(va) and not _is_numeric(va):
                    text_in_a += 1
                for c in other_c:
                    v = adapter.cell(r, c)
                    if not _is_empty(v):
                        total_rest += 1
                        if _is_numeric(v):
                            numeric_in_rest += 1

            if text_in_a < 2:
                continue
            if total_rest == 0 or numeric_in_rest / total_rest < 0.5:
                continue

            data_rows = [r for r in data_candidates if r not in used_rows]
            if not data_rows:
                continue

            raw_headers = [_to_str(adapter.cell(header_row, c)) for c in vis_cols]
            deduped_hdr = _dedupe_headers(raw_headers)
            col_map: dict[int, str] = {c: deduped_hdr[i] for i, c in enumerate(vis_cols)}

            # FIX v19: используем _build_rows
            rows_out, col_values = self._build_rows(adapter, data_rows, col_map)

            if not rows_out:
                continue

            tbl_num = len(results) + 1
            results.append({
                "sheet": adapter.name,
                "name": f"{adapter.name} / Вертикальная таблица {tbl_num}",
                "source": TABLE_SOURCE_VERTICAL,
                "header_row": header_row,
                "data_start": data_rows[0],
                "data_end": data_rows[-1],
                "columns": [
                    {
                        "letter": get_column_letter(c),
                        "name": col_map[c],
                        "type": _detect_dtype(col_values[c]),
                    }
                    for c in vis_cols
                ],
                "rows": rows_out,
            })
            used_rows.update(range(header_row, data_rows[-1] + 1))

        return results

    # -- Источник 5: таблицы без заголовка --------------------------------------

    def _extract_headerless(self, adapter: SheetAdapter, used_rows: set[int]) -> list[dict]:
        vis_rows, vis_cols = self._visible(adapter)
        if not vis_rows or not vis_cols:
            return []

        tables: list[dict] = []
        table_counter = 0
        i = 0

        while i < len(vis_rows):
            r = vis_rows[i]
            vals = [adapter.cell(r, c) for c in vis_cols]

            if all(_is_empty(v) for v in vals):
                i += 1
                continue

            if r in used_rows:
                i += 1
                continue

            block_rows: list[int] = []
            empty_streak = 0
            for j in range(i, len(vis_rows)):
                rj = vis_rows[j]
                if rj in used_rows:
                    break
                rv = [adapter.cell(rj, c) for c in vis_cols]
                non_e = sum(1 for v in rv if not _is_empty(v))
                if non_e == 0:
                    empty_streak += 1
                    if empty_streak >= self.max_empty_streak:
                        break
                    continue
                empty_streak = 0
                block_rows.append(rj)

            if len(block_rows) < 2:
                i += len(block_rows) + 1
                continue

            col_map: dict[int, str] = {c: get_column_letter(c) for c in vis_cols}

            # FIX v19: используем _build_rows
            rows_out, col_values = self._build_rows(adapter, block_rows, col_map)

            if rows_out:
                table_counter += 1
                tables.append({
                    "sheet": adapter.name,
                    "name": f"{adapter.name} / Беззаголовочная {table_counter}",
                    "source": TABLE_SOURCE_HEADERLESS,
                    "header_row": block_rows[0],
                    "data_start": block_rows[0],
                    "data_end": block_rows[-1],
                    "columns": [
                        {
                            "letter": get_column_letter(c),
                            "name": col_map[c],
                            "type": _detect_dtype(col_values[c]),
                        }
                        for c in vis_cols
                    ],
                    "rows": rows_out,
                })
                used_rows.update(range(block_rows[0], block_rows[-1] + 1))

            i += len(block_rows) + 1

        return tables

    # -- parse_sheet ------------------------------------------------------------

    def parse_sheet(
        self,
        adapter: SheetAdapter,
        wb=None,
        all_adapters: Optional[dict[str, SheetAdapter]] = None,
    ) -> list[dict]:
        self._vis_cache.clear()

        all_tables: list[dict] = []
        used_rows: set[int] = set()

        # 1. Встроенные таблицы
        native = self._extract_native_tables(adapter)
        for t in native:
            used_rows.update(range(t["header_row"], t["data_end"] + 1))
        all_tables.extend(native)

        # 2. Именованные диапазоны (ищем всегда)
        named: list[dict] = []
        if wb is not None and all_adapters:
            named = [t for t in self._extract_named_ranges_from_wb(wb, all_adapters)
                     if t["sheet"] == adapter.name]
        else:
            named = self._extract_named_ranges_from_adapter(adapter)

        for t in named:
            span = set(range(t["header_row"], t["data_end"] + 1))
            if not (span & used_rows):
                used_rows.update(span)
                all_tables.append(t)

        # 3. Эвристика — FIX v19: запускается ВСЕГДА, не только при пустом all_tables
        heuristic = self._extract_heuristic(adapter, used_rows)
        for t in heuristic:
            used_rows.update(range(t["header_row"], t["data_end"] + 1))
        all_tables.extend(heuristic)

        # 4. Вертикальные таблицы
        all_tables.extend(self._extract_vertical(adapter, used_rows))

        # 5. Беззаголовочные таблицы
        all_tables.extend(self._extract_headerless(adapter, used_rows))

        return all_tables

    # -- parse_file -------------------------------------------------------------

    def parse_file(
        self,
        filepath: str,
        output_path: Optional[str] = None,
        fmt: str = "json",
        only_sheet: Optional[str] = None,
        streaming: bool = False,
    ) -> dict:
        print(f"\n{'=' * 72}")
        print(f"  {os.path.basename(filepath)}")
        print(f"{'=' * 72}")

        ext = os.path.splitext(filepath)[1].lower()
        adapters = load_sheets(filepath, only_sheet)

        wb = None
        if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            first_opx = next(
                (a for a in adapters if isinstance(a, OpenpyxlAdapter)), None
            )
            if first_opx is not None:
                wb = first_opx._ws.parent
        adapters_map = {a.name: a for a in adapters}

        writer: Optional[StreamingWriter] = None
        if streaming and output_path:
            file_meta = {"file": os.path.basename(filepath), "format": ext.lstrip(".")}
            writer = StreamingWriter(output_path, fmt, file_meta)
        elif streaming and not output_path:
            warnings.warn(
                "--stream требует --out-dir или явного output_path. "
                "Данные не будут сохранены на диск.",
                UserWarning, stacklevel=2,
            )

        all_tables: list[dict] = []
        stream_sources: dict[str, int] = {
            TABLE_SOURCE_NATIVE: 0,
            TABLE_SOURCE_NAMED: 0,
            TABLE_SOURCE_HEURISTIC: 0,
            TABLE_SOURCE_VERTICAL: 0,
            TABLE_SOURCE_HEADERLESS: 0,
        }
        iter_list = (
            _tqdm(adapters, desc="Листы", unit="лист")
            if HAS_TQDM and len(adapters) > 3
            else adapters
        )

        parse_error = None
        try:
            for adapter in iter_list:
                tables = self.parse_sheet(adapter, wb=wb, all_adapters=adapters_map)

                src_counts: dict[str, int] = {}
                for t in tables:
                    src = t["source"]
                    src_counts[src] = src_counts.get(src, 0) + 1
                    if src in stream_sources:
                        stream_sources[src] += 1
                    if writer:
                        writer.write_table(t)

                summary = ", ".join(f"{k}:{v}" for k, v in src_counts.items()) or "-"
                print(f"  '{adapter.name}' "
                      f"({adapter.max_row}x{adapter.max_col}) "
                      f"-> {len(tables)} таблиц [{summary}]")

                if not streaming:
                    all_tables.extend(tables)
        except Exception as e:
            parse_error = e
        finally:
            if writer:
                try:
                    if parse_error is None:
                        writer.close()
                    else:
                        # При ошибке закрываем файл без финализации JSON
                        if writer._json_fh:
                            writer._json_fh.close()
                            writer._json_fh = None
                        if writer._jsonl_fh:
                            writer._jsonl_fh.close()
                            writer._jsonl_fh = None
                except Exception:
                    pass  # не маскируем parse_error

        if parse_error is not None:
            raise parse_error

        if writer:
            n_tables, n_rows = writer.stats
        else:
            n_tables = len(all_tables)
            n_rows = sum(len(t["rows"]) for t in all_tables)

        result = {
            "file": os.path.basename(filepath),
            "format": ext.lstrip("."),
            "sheets": len(adapters),
            "tables": n_tables,
            "total_rows": n_rows,
            "sources": stream_sources,
            "tables_data": all_tables,
        }

        if output_path and not streaming:
            _write_output(result, output_path, fmt)
            print(f"  -> {output_path}")

        print(f"\n{'=' * 72}")
        print(
            f"  {n_tables} таблиц | {n_rows} строк | "
            f"native:{result['sources'][TABLE_SOURCE_NATIVE]} "
            f"named:{result['sources'][TABLE_SOURCE_NAMED]} "
            f"heuristic:{result['sources'][TABLE_SOURCE_HEURISTIC]} "
            f"vertical:{result['sources'][TABLE_SOURCE_VERTICAL]} "
            f"headerless:{result['sources'][TABLE_SOURCE_HEADERLESS]}"
            + (" | streaming" if streaming else "")
        )
        print(f"{'=' * 72}\n")

        return result


# ==============================================================================
# Запись (не-потоковый режим)
# ==============================================================================

def _write_output(result: dict, path: str, fmt: str) -> None:
    if fmt == "json":
        with open(path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
    elif fmt == "jsonl":
        with open(path, "w", encoding="utf-8") as f:
            for table in result["tables_data"]:
                for row in table["rows"]:
                    record = {"_sheet": table["sheet"], "_table": table["name"], **row}
                    f.write(json.dumps(record, ensure_ascii=False, default=str) + "\n")
    elif fmt == "csv":
        os.makedirs(path, exist_ok=True)
        for idx, table in enumerate(result["tables_data"]):
            rows = table.get("rows", [])
            if not rows:
                continue
            safe = (
                table["name"]
                .replace("/", "_").replace("\\", "_")
                .replace(":", "_").replace(" ", "_")[:60]
            )
            csv_path = os.path.join(path, f"{idx + 1:02d}_{safe}.csv")
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.DictWriter(f, fieldnames=list(rows[0].keys()), extrasaction="ignore")
                w.writeheader()
                w.writerows(rows)
    else:
        raise ValueError(f"Неизвестный формат: {fmt}")


# ==============================================================================
# CLI
# ==============================================================================

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Excel Universal Parser v19",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    ap.add_argument("files", nargs="+",
                    help="Файлы .xlsx .xlsm .xls .xlsb .csv")
    ap.add_argument("--out-dir", default=None,
                    help="Директория для результатов")
    ap.add_argument("--format", choices=["json", "jsonl", "csv"], default="json",
                    help="Формат вывода (default: json)")
    ap.add_argument("--sheet", default=None,
                    help="Обрабатывать только этот лист")
    ap.add_argument("--header-threshold", type=float, default=0.4,
                    help="Score-порог для заголовка [0..1] (default: 0.4)")
    ap.add_argument("--min-data-cells", type=int, default=2,
                    help="Мин. непустых ячеек для строки данных (default: 2)")
    # FIX v19: заменён --include-hidden (не работал) на --skip-hidden
    ap.add_argument("--skip-hidden", action="store_true",
                    help="Пропускать скрытые строки и колонки")
    ap.add_argument("--stream", action="store_true",
                    help="Потоковая запись -- не накапливать данные в RAM")
    args = ap.parse_args()

    parser = ExcelParser(
        header_threshold=args.header_threshold,
        skip_hidden=args.skip_hidden,
        min_data_cells=args.min_data_cells,
    )

    for fp in args.files:
        if not os.path.exists(fp):
            print(f"  Файл не найден: {fp}", file=sys.stderr)
            continue

        out_dir = args.out_dir or os.path.dirname(os.path.abspath(fp))
        stem = os.path.splitext(os.path.basename(fp))[0].replace(".", "_")

        if args.format == "csv":
            out_path = os.path.join(out_dir, stem + "_tables")
        else:
            out_path = os.path.join(out_dir, stem + "_parsed." + args.format)

        try:
            parser.parse_file(
                fp,
                output_path=out_path,
                fmt=args.format,
                only_sheet=args.sheet,
                streaming=args.stream,
            )
        except Exception as exc:
            print(f"  Ошибка: {fp}\n    {exc}", file=sys.stderr)
            traceback.print_exc()


if __name__ == "__main__":
    main()
